from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from src.models import Metric


def extract_excel(file_path: Path) -> list[Metric]:
    """Extract metrics from an Excel file.

    Expects a header row with columns that map to metric fields.
    Recognized header names (case-insensitive):
        metric_name / name / metric
        metric_value / value / amount
        metric_unit / unit / currency
        period / date / month
        category / type / group
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return []

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]

    col_map = _build_column_map(headers)
    if "metric_name" not in col_map or "metric_value" not in col_map:
        # Fallback: treat first column as name, second as value
        if len(headers) >= 2:
            col_map = {"metric_name": 0, "metric_value": 1}
        else:
            return []

    metrics: list[Metric] = []
    for row in rows[1:]:
        name = _cell(row, col_map, "metric_name")
        value_raw = _cell(row, col_map, "metric_value")
        if not name or value_raw is None:
            continue
        try:
            value = float(value_raw)
        except (TypeError, ValueError):
            continue

        metrics.append(
            Metric(
                metric_name=str(name),
                metric_value=value,
                metric_unit=str(_cell(row, col_map, "metric_unit") or ""),
                period=str(_cell(row, col_map, "period") or ""),
                category=str(_cell(row, col_map, "category") or ""),
            )
        )

    return metrics


def _build_column_map(headers: list[str]) -> dict[str, int]:
    aliases: dict[str, list[str]] = {
        "metric_name": ["metric_name", "name", "metric", "kpi"],
        "metric_value": ["metric_value", "value", "amount", "total"],
        "metric_unit": ["metric_unit", "unit", "currency"],
        "period": ["period", "date", "month", "year"],
        "category": ["category", "type", "group", "department"],
    }
    col_map: dict[str, int] = {}
    for field, names in aliases.items():
        for i, h in enumerate(headers):
            if h in names:
                col_map[field] = i
                break
    return col_map


def _cell(row: tuple, col_map: dict[str, int], field: str):
    idx = col_map.get(field)
    if idx is None or idx >= len(row):
        return None
    return row[idx]
